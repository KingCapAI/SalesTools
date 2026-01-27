"""
API routes for quote calculations.
"""

from io import BytesIO
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ..schemas.quote import (
    DomesticQuoteRequest,
    OverseasQuoteRequest,
    DomesticQuoteResponse,
    OverseasQuoteResponse,
    QuoteOptionsResponse,
    QuoteSheetExportRequest,
)
from ..services.pricing_service import (
    calculate_domestic_quote,
    calculate_overseas_quote,
)
from ..data.pricing import (
    DOMESTIC_QUANTITY_BREAKS,
    OVERSEAS_QUANTITY_BREAKS,
    DOMESTIC_STYLES,
    DOMESTIC_FRONT_DECORATION_METHODS,
    DOMESTIC_ADDITIONAL_DECORATION_METHODS,
    DOMESTIC_RUSH_FEES,
    OVERSEAS_HAT_TYPES,
    OVERSEAS_DECORATION_METHODS,
    OVERSEAS_DESIGN_ADDONS,
    OVERSEAS_ACCESSORIES,
    OVERSEAS_SHIPPING,
)

router = APIRouter(prefix="/quotes", tags=["quotes"])


@router.get("/options", response_model=QuoteOptionsResponse)
async def get_quote_options():
    """Get all available options for quote calculations."""
    return {
        "domestic": {
            "quantity_breaks": DOMESTIC_QUANTITY_BREAKS,
            "styles": [
                {
                    "style_number": style_num,
                    "name": info["name"],
                    "collection": info["collection"],
                }
                for style_num, info in DOMESTIC_STYLES.items()
            ],
            "front_decoration_methods": DOMESTIC_FRONT_DECORATION_METHODS,
            "additional_decoration_methods": DOMESTIC_ADDITIONAL_DECORATION_METHODS,
            "shipping_speeds": list(DOMESTIC_RUSH_FEES.keys()),
        },
        "overseas": {
            "quantity_breaks": OVERSEAS_QUANTITY_BREAKS,
            "hat_types": list(OVERSEAS_HAT_TYPES.keys()),
            "decoration_methods": OVERSEAS_DECORATION_METHODS,
            "design_addons": list(OVERSEAS_DESIGN_ADDONS.keys()),
            "accessories": list(OVERSEAS_ACCESSORIES.keys()),
            "shipping_methods": list(OVERSEAS_SHIPPING.keys()),
        },
    }


@router.post("/domestic", response_model=DomesticQuoteResponse)
async def calculate_domestic(request: DomesticQuoteRequest):
    """Calculate a domestic quote."""
    try:
        result = calculate_domestic_quote(
            style_number=request.style_number,
            quantity=request.quantity,
            front_decoration=request.front_decoration,
            left_decoration=request.left_decoration,
            right_decoration=request.right_decoration,
            back_decoration=request.back_decoration,
            shipping_speed=request.shipping_speed,
            include_rope=request.include_rope,
            num_dst_files=request.num_dst_files,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/overseas", response_model=OverseasQuoteResponse)
async def calculate_overseas(request: OverseasQuoteRequest):
    """Calculate an overseas quote."""
    try:
        result = calculate_overseas_quote(
            hat_type=request.hat_type,
            quantity=request.quantity,
            front_decoration=request.front_decoration,
            left_decoration=request.left_decoration,
            right_decoration=request.right_decoration,
            back_decoration=request.back_decoration,
            visor_decoration=request.visor_decoration,
            design_addons=request.design_addons,
            accessories=request.accessories,
            shipping_method=request.shipping_method,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


def style_header_cell(cell):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_currency_cell(cell, value):
    """Apply currency formatting to a cell."""
    if value is None:
        cell.value = "Does not meet MOQ"
        cell.font = Font(italic=True, color="888888")
    else:
        cell.value = value
        cell.number_format = '"$"#,##0.00'


@router.post("/domestic/export")
async def export_domestic_quote(request: DomesticQuoteRequest):
    """Export a domestic quote to Excel."""
    try:
        result = calculate_domestic_quote(
            style_number=request.style_number,
            quantity=request.quantity,
            front_decoration=request.front_decoration,
            left_decoration=request.left_decoration,
            right_decoration=request.right_decoration,
            back_decoration=request.back_decoration,
            shipping_speed=request.shipping_speed,
            include_rope=request.include_rope,
            num_dst_files=request.num_dst_files,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Domestic Quote"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "King Cap - Domestic Quote"
    ws["A1"].font = Font(bold=True, size=14)

    # Quote details
    row = 3
    details = []
    if request.design_number:
        details.append(("Design #:", request.design_number))
    details.extend([
        ("Style:", f"{result['style_number']} - {result['style_name']}"),
        ("Collection:", result["collection"]),
        ("Quantity:", f"{result['quantity']:,}"),
        ("Shipping:", result["shipping_speed"]),
    ])
    for label, value in details:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Price table header
    row += 1
    headers = ["Line Item", "Per Piece", "Qty", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        style_header_cell(cell)

    # Get the applicable price break
    pb = result["price_breaks"][-1] if result["price_breaks"] else None
    if not pb:
        raise HTTPException(status_code=400, detail="No price breaks available")

    # Helper to check valid decoration
    def is_valid(val):
        return isinstance(val, str) and val.strip() and val != "0"

    # Price rows
    row += 1
    line_items = [("Blank Hat", pb["blank_price"], result["quantity"])]

    if is_valid(result.get("front_decoration")):
        line_items.append((f"Front Decoration ({result['front_decoration']})", pb["front_decoration_price"], result["quantity"]))
    if is_valid(result.get("left_decoration")):
        line_items.append((f"Left Decoration ({result['left_decoration']})", pb["left_decoration_price"], result["quantity"]))
    if is_valid(result.get("right_decoration")):
        line_items.append((f"Right Decoration ({result['right_decoration']})", pb["right_decoration_price"], result["quantity"]))
    if is_valid(result.get("back_decoration")):
        line_items.append((f"Back Decoration ({result['back_decoration']})", pb["back_decoration_price"], result["quantity"]))
    if pb.get("rush_fee", 0) > 0:
        line_items.append(("Rush Fee", pb["rush_fee"], result["quantity"]))
    if result.get("include_rope") and pb.get("rope_price", 0) > 0:
        line_items.append(("Rope", pb["rope_price"], result["quantity"]))
    if pb.get("digitizing_fee", 0) > 0:
        line_items.append(("Digitizing Fee", pb["digitizing_fee"], 1))

    for label, per_piece, qty in line_items:
        ws.cell(row=row, column=1, value=label)
        style_currency_cell(ws.cell(row=row, column=2), per_piece)
        ws.cell(row=row, column=3, value=qty)
        total = per_piece * qty if per_piece else 0
        style_currency_cell(ws.cell(row=row, column=4), total)
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=1).font = Font(bold=True)
    style_currency_cell(ws.cell(row=row, column=2), pb["per_piece_price"])
    ws.cell(row=row, column=3, value=result["quantity"])
    style_currency_cell(ws.cell(row=row, column=4), pb["total"])
    ws.cell(row=row, column=4).font = Font(bold=True)

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    design_part = f"_{request.design_number}" if request.design_number else ""
    filename = f"domestic_quote{design_part}_{result['style_number']}_{result['quantity']}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/overseas/export")
async def export_overseas_quote(request: OverseasQuoteRequest):
    """Export an overseas quote to Excel."""
    try:
        result = calculate_overseas_quote(
            hat_type=request.hat_type,
            quantity=request.quantity,
            front_decoration=request.front_decoration,
            left_decoration=request.left_decoration,
            right_decoration=request.right_decoration,
            back_decoration=request.back_decoration,
            visor_decoration=request.visor_decoration,
            design_addons=request.design_addons,
            accessories=request.accessories,
            shipping_method=request.shipping_method,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Overseas Quote"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "King Cap - Overseas Quote"
    ws["A1"].font = Font(bold=True, size=14)

    # Quote details
    row = 3
    details = []
    if request.design_number:
        details.append(("Design #:", request.design_number))
    details.extend([
        ("Hat Type:", result["hat_type"]),
        ("Shipping:", result["shipping_method"]),
    ])

    # Add decoration details
    def is_valid(val):
        return isinstance(val, str) and val.strip() and val != "0"

    if is_valid(result.get("front_decoration")):
        details.append(("Front:", result["front_decoration"]))
    if is_valid(result.get("left_decoration")):
        details.append(("Left:", result["left_decoration"]))
    if is_valid(result.get("right_decoration")):
        details.append(("Right:", result["right_decoration"]))
    if is_valid(result.get("back_decoration")):
        details.append(("Back:", result["back_decoration"]))
    if is_valid(result.get("visor_decoration")):
        details.append(("Visor:", result["visor_decoration"]))
    if result.get("design_addons"):
        details.append(("Add-ons:", ", ".join(result["design_addons"])))
    if result.get("accessories"):
        details.append(("Accessories:", ", ".join(result["accessories"])))

    for label, value in details:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Price table header
    row += 1
    price_breaks = result["price_breaks"]

    # Header row: Line Item, then quantity breaks
    ws.cell(row=row, column=1, value="Line Item")
    style_header_cell(ws.cell(row=row, column=1))

    for col, pb in enumerate(price_breaks, 2):
        cell = ws.cell(row=row, column=col, value=f"{pb['quantity_break']:,}+")
        style_header_cell(cell)

    # Hat row
    row += 1
    ws.cell(row=row, column=1, value="Hat")
    for col, pb in enumerate(price_breaks, 2):
        if pb["per_piece_price"] is None:
            style_currency_cell(ws.cell(row=row, column=col), None)
        else:
            hat_cost = (
                (pb["blank_price"] or 0) +
                (pb["front_decoration_price"] or 0) +
                (pb["left_decoration_price"] or 0) +
                (pb["right_decoration_price"] or 0) +
                (pb["back_decoration_price"] or 0) +
                (pb["visor_decoration_price"] or 0) +
                (pb["addons_price"] or 0) +
                (pb["accessories_price"] or 0)
            )
            style_currency_cell(ws.cell(row=row, column=col), hat_cost)

    # Shipping row
    row += 1
    ws.cell(row=row, column=1, value="Shipping")
    for col, pb in enumerate(price_breaks, 2):
        if pb["per_piece_price"] is None:
            style_currency_cell(ws.cell(row=row, column=col), None)
        else:
            style_currency_cell(ws.cell(row=row, column=col), pb["shipping_price"] or 0)

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    for col in range(2, len(price_breaks) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Footer note
    row += 2
    ws.cell(row=row, column=1, value="* All prices shown are per piece at each quantity break")
    ws.cell(row=row, column=1).font = Font(italic=True, color="666666")

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    design_part = f"_{request.design_number}" if request.design_number else ""
    filename = f"overseas_quote{design_part}_{result['hat_type'].lower().replace(' ', '_')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/sheet/export")
async def export_quote_sheet(request: QuoteSheetExportRequest):
    """Export a combined quote sheet with multiple designs to Excel."""
    if not request.quotes:
        raise HTTPException(status_code=400, detail="No quotes provided")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Sheet"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "King Cap - Combined Quote Sheet"
    ws["A1"].font = Font(bold=True, size=14)

    current_row = 3

    for quote_item in request.quotes:
        try:
            if quote_item.type == "domestic":
                # Parse request
                req = quote_item.request
                result = calculate_domestic_quote(
                    style_number=req.get("style_number"),
                    quantity=req.get("quantity"),
                    front_decoration=req.get("front_decoration"),
                    left_decoration=req.get("left_decoration"),
                    right_decoration=req.get("right_decoration"),
                    back_decoration=req.get("back_decoration"),
                    shipping_speed=req.get("shipping_speed", "Standard (5-7 Production Days)"),
                    include_rope=req.get("include_rope", False),
                    num_dst_files=req.get("num_dst_files", 1),
                )

                # Design header
                ws.cell(row=current_row, column=1, value=f"Design: {quote_item.design_number}")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                current_row += 1

                # Details
                ws.cell(row=current_row, column=1, value="Type:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value="Domestic")
                current_row += 1

                ws.cell(row=current_row, column=1, value="Style:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=f"{result['style_number']} - {result['style_name']}")
                current_row += 1

                ws.cell(row=current_row, column=1, value="Quantity:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=f"{result['quantity']:,}")
                current_row += 1

                # Price table header
                current_row += 1
                headers = ["Line Item", "Per Piece", "Qty", "Total"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    style_header_cell(cell)

                # Get the applicable price break
                pb = result["price_breaks"][-1] if result["price_breaks"] else None
                if pb:
                    # Helper to check valid decoration
                    def is_valid(val):
                        return isinstance(val, str) and val.strip() and val != "0"

                    # Price rows
                    current_row += 1
                    line_items = [("Blank Hat", pb["blank_price"], result["quantity"])]

                    if is_valid(result.get("front_decoration")):
                        line_items.append((f"Front Decoration ({result['front_decoration']})", pb["front_decoration_price"], result["quantity"]))
                    if is_valid(result.get("left_decoration")):
                        line_items.append((f"Left Decoration ({result['left_decoration']})", pb["left_decoration_price"], result["quantity"]))
                    if is_valid(result.get("right_decoration")):
                        line_items.append((f"Right Decoration ({result['right_decoration']})", pb["right_decoration_price"], result["quantity"]))
                    if is_valid(result.get("back_decoration")):
                        line_items.append((f"Back Decoration ({result['back_decoration']})", pb["back_decoration_price"], result["quantity"]))
                    if pb.get("rush_fee", 0) > 0:
                        line_items.append(("Rush Fee", pb["rush_fee"], result["quantity"]))
                    if result.get("include_rope") and pb.get("rope_price", 0) > 0:
                        line_items.append(("Rope", pb["rope_price"], result["quantity"]))
                    if pb.get("digitizing_fee", 0) > 0:
                        line_items.append(("Digitizing Fee", pb["digitizing_fee"], 1))

                    for label, per_piece, qty in line_items:
                        ws.cell(row=current_row, column=1, value=label)
                        style_currency_cell(ws.cell(row=current_row, column=2), per_piece)
                        ws.cell(row=current_row, column=3, value=qty)
                        total = per_piece * qty if per_piece else 0
                        style_currency_cell(ws.cell(row=current_row, column=4), total)
                        current_row += 1

                    # Total row
                    ws.cell(row=current_row, column=1, value="Total")
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    style_currency_cell(ws.cell(row=current_row, column=2), pb["per_piece_price"])
                    ws.cell(row=current_row, column=3, value=result["quantity"])
                    style_currency_cell(ws.cell(row=current_row, column=4), pb["total"])
                    ws.cell(row=current_row, column=4).font = Font(bold=True)
                    current_row += 1

            else:  # overseas
                req = quote_item.request
                result = calculate_overseas_quote(
                    hat_type=req.get("hat_type"),
                    quantity=req.get("quantity", 5040),
                    front_decoration=req.get("front_decoration"),
                    left_decoration=req.get("left_decoration"),
                    right_decoration=req.get("right_decoration"),
                    back_decoration=req.get("back_decoration"),
                    visor_decoration=req.get("visor_decoration"),
                    design_addons=req.get("design_addons"),
                    accessories=req.get("accessories"),
                    shipping_method=req.get("shipping_method", "FOB CA"),
                )

                # Design header
                ws.cell(row=current_row, column=1, value=f"Design: {quote_item.design_number}")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(result["price_breaks"]) + 1)
                current_row += 1

                # Details
                ws.cell(row=current_row, column=1, value="Type:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value="Overseas")
                current_row += 1

                ws.cell(row=current_row, column=1, value="Hat Type:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=result["hat_type"])
                current_row += 1

                ws.cell(row=current_row, column=1, value="Shipping:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=result["shipping_method"])
                current_row += 1

                # Price table header
                current_row += 1
                price_breaks = result["price_breaks"]

                ws.cell(row=current_row, column=1, value="Line Item")
                style_header_cell(ws.cell(row=current_row, column=1))

                for col, pb in enumerate(price_breaks, 2):
                    cell = ws.cell(row=current_row, column=col, value=f"{pb['quantity_break']:,}+")
                    style_header_cell(cell)

                # Hat row
                current_row += 1
                ws.cell(row=current_row, column=1, value="Hat")
                for col, pb in enumerate(price_breaks, 2):
                    if pb["per_piece_price"] is None:
                        style_currency_cell(ws.cell(row=current_row, column=col), None)
                    else:
                        hat_cost = (
                            (pb["blank_price"] or 0) +
                            (pb["front_decoration_price"] or 0) +
                            (pb["left_decoration_price"] or 0) +
                            (pb["right_decoration_price"] or 0) +
                            (pb["back_decoration_price"] or 0) +
                            (pb["visor_decoration_price"] or 0) +
                            (pb["addons_price"] or 0) +
                            (pb["accessories_price"] or 0)
                        )
                        style_currency_cell(ws.cell(row=current_row, column=col), hat_cost)

                # Shipping row
                current_row += 1
                ws.cell(row=current_row, column=1, value="Shipping")
                for col, pb in enumerate(price_breaks, 2):
                    if pb["per_piece_price"] is None:
                        style_currency_cell(ws.cell(row=current_row, column=col), None)
                    else:
                        style_currency_cell(ws.cell(row=current_row, column=col), pb["shipping_price"] or 0)

                current_row += 1

        except Exception as e:
            # If a quote fails, add error note and continue
            ws.cell(row=current_row, column=1, value=f"Error processing {quote_item.design_number}: {str(e)}")
            ws.cell(row=current_row, column=1).font = Font(color="FF0000")
            current_row += 1

        # Add spacing between designs
        current_row += 2

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15
    for col in range(5, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quote_sheet.xlsx"}
    )
