"""Design quote management routes."""

import json
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from ..database import get_db
from ..models import Design, DesignQuote
from ..schemas.design_quote import (
    DesignQuoteCreate,
    DesignQuoteUpdate,
    DesignQuoteResponse,
)
from ..services.pricing_service import calculate_domestic_quote, calculate_overseas_quote
from ..utils.dependencies import require_auth

router = APIRouter(prefix="/designs", tags=["Design Quotes"])


def _parse_json_field(value: Optional[str]) -> Optional[list]:
    """Parse a JSON string field to list."""
    if not value:
        return None
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return None


def _serialize_json_field(value: Optional[list]) -> Optional[str]:
    """Serialize a list to JSON string."""
    if not value:
        return None
    return json.dumps(value)


def _quote_to_response(quote: DesignQuote) -> dict:
    """Convert DesignQuote model to response dict with parsed JSON fields."""
    return {
        "id": quote.id,
        "design_id": quote.design_id,
        "quote_type": quote.quote_type,
        "quantity": quote.quantity,
        "front_decoration": quote.front_decoration,
        "left_decoration": quote.left_decoration,
        "right_decoration": quote.right_decoration,
        "back_decoration": quote.back_decoration,
        "style_number": quote.style_number,
        "shipping_speed": quote.shipping_speed,
        "include_rope": quote.include_rope,
        "num_dst_files": quote.num_dst_files,
        "hat_type": quote.hat_type,
        "visor_decoration": quote.visor_decoration,
        "design_addons": _parse_json_field(quote.design_addons),
        "accessories": _parse_json_field(quote.accessories),
        "shipping_method": quote.shipping_method,
        "cached_price_breaks": _parse_json_field(quote.cached_price_breaks),
        "cached_total": quote.cached_total / 100 if quote.cached_total else None,
        "cached_per_piece": quote.cached_per_piece / 100 if quote.cached_per_piece else None,
        "created_at": quote.created_at,
        "updated_at": quote.updated_at,
    }


@router.get("/{design_id}/quote", response_model=Optional[DesignQuoteResponse])
async def get_design_quote(
    design_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_auth),
):
    """Get the quote for a design (if exists)."""
    quote = db.query(DesignQuote).filter(DesignQuote.design_id == design_id).first()
    if not quote:
        return None
    return _quote_to_response(quote)


@router.post("/{design_id}/quote", response_model=DesignQuoteResponse)
async def create_design_quote(
    design_id: str,
    quote_data: DesignQuoteCreate,
    db: Session = Depends(get_db),
    user=Depends(require_auth),
):
    """Create or replace quote for a design."""
    # Verify design exists
    design = db.query(Design).filter(Design.id == design_id).first()
    if not design:
        raise HTTPException(status_code=404, detail="Design not found")

    # Delete existing quote if any
    existing = db.query(DesignQuote).filter(DesignQuote.design_id == design_id).first()
    if existing:
        db.delete(existing)
        db.flush()

    # Calculate quote
    try:
        if quote_data.quote_type == "domestic":
            if not quote_data.style_number:
                raise HTTPException(status_code=400, detail="style_number is required for domestic quotes")
            result = calculate_domestic_quote(
                style_number=quote_data.style_number,
                quantity=quote_data.quantity,
                front_decoration=quote_data.front_decoration,
                left_decoration=quote_data.left_decoration,
                right_decoration=quote_data.right_decoration,
                back_decoration=quote_data.back_decoration,
                shipping_speed=quote_data.shipping_speed or "Standard (5-7 Production Days)",
                include_rope=quote_data.include_rope or False,
                num_dst_files=quote_data.num_dst_files or 1,
            )
            price_breaks = result["price_breaks"]
            applicable_break = price_breaks[-1] if price_breaks else None
        else:
            if not quote_data.hat_type:
                raise HTTPException(status_code=400, detail="hat_type is required for overseas quotes")
            result = calculate_overseas_quote(
                hat_type=quote_data.hat_type,
                quantity=quote_data.quantity,
                front_decoration=quote_data.front_decoration,
                left_decoration=quote_data.left_decoration,
                right_decoration=quote_data.right_decoration,
                back_decoration=quote_data.back_decoration,
                visor_decoration=quote_data.visor_decoration,
                design_addons=quote_data.design_addons,
                accessories=quote_data.accessories,
                shipping_method=quote_data.shipping_method or "FOB CA",
            )
            price_breaks = result["price_breaks"]
            # Find the applicable price break for the requested quantity
            applicable_break = None
            for pb in reversed(price_breaks):
                if pb["quantity_break"] <= quote_data.quantity and pb.get("per_piece_price"):
                    applicable_break = pb
                    break
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Create new quote
    quote = DesignQuote(
        design_id=design_id,
        quote_type=quote_data.quote_type,
        quantity=quote_data.quantity,
        front_decoration=quote_data.front_decoration,
        left_decoration=quote_data.left_decoration,
        right_decoration=quote_data.right_decoration,
        back_decoration=quote_data.back_decoration,
        style_number=quote_data.style_number,
        shipping_speed=quote_data.shipping_speed,
        include_rope=quote_data.include_rope,
        num_dst_files=quote_data.num_dst_files,
        hat_type=quote_data.hat_type,
        visor_decoration=quote_data.visor_decoration,
        design_addons=_serialize_json_field(quote_data.design_addons),
        accessories=_serialize_json_field(quote_data.accessories),
        shipping_method=quote_data.shipping_method,
        cached_price_breaks=json.dumps(price_breaks),
        cached_total=int(applicable_break["total"] * 100) if applicable_break and applicable_break.get("total") else None,
        cached_per_piece=int(applicable_break["per_piece_price"] * 100) if applicable_break and applicable_break.get("per_piece_price") else None,
        created_by_id=str(user.id),
    )

    db.add(quote)
    db.commit()
    db.refresh(quote)
    return _quote_to_response(quote)


@router.patch("/{design_id}/quote", response_model=DesignQuoteResponse)
async def update_design_quote(
    design_id: str,
    update_data: DesignQuoteUpdate,
    db: Session = Depends(get_db),
    user=Depends(require_auth),
):
    """Update existing quote and recalculate."""
    quote = db.query(DesignQuote).filter(DesignQuote.design_id == design_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    # Update fields
    update_dict = update_data.model_dump(exclude_unset=True)
    for field, value in update_dict.items():
        if field in ["design_addons", "accessories"]:
            setattr(quote, field, _serialize_json_field(value))
        else:
            setattr(quote, field, value)

    # Recalculate
    try:
        if quote.quote_type == "domestic":
            result = calculate_domestic_quote(
                style_number=quote.style_number,
                quantity=quote.quantity,
                front_decoration=quote.front_decoration,
                left_decoration=quote.left_decoration,
                right_decoration=quote.right_decoration,
                back_decoration=quote.back_decoration,
                shipping_speed=quote.shipping_speed or "Standard (5-7 Production Days)",
                include_rope=quote.include_rope or False,
                num_dst_files=quote.num_dst_files or 1,
            )
            price_breaks = result["price_breaks"]
            applicable_break = price_breaks[-1] if price_breaks else None
        else:
            result = calculate_overseas_quote(
                hat_type=quote.hat_type,
                quantity=quote.quantity,
                front_decoration=quote.front_decoration,
                left_decoration=quote.left_decoration,
                right_decoration=quote.right_decoration,
                back_decoration=quote.back_decoration,
                visor_decoration=quote.visor_decoration,
                design_addons=_parse_json_field(quote.design_addons),
                accessories=_parse_json_field(quote.accessories),
                shipping_method=quote.shipping_method or "FOB CA",
            )
            price_breaks = result["price_breaks"]
            applicable_break = None
            for pb in reversed(price_breaks):
                if pb["quantity_break"] <= quote.quantity and pb.get("per_piece_price"):
                    applicable_break = pb
                    break
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    quote.cached_price_breaks = json.dumps(price_breaks)
    quote.cached_total = int(applicable_break["total"] * 100) if applicable_break and applicable_break.get("total") else None
    quote.cached_per_piece = int(applicable_break["per_piece_price"] * 100) if applicable_break and applicable_break.get("per_piece_price") else None

    db.commit()
    db.refresh(quote)
    return _quote_to_response(quote)


@router.delete("/{design_id}/quote")
async def delete_design_quote(
    design_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_auth),
):
    """Delete quote from design."""
    quote = db.query(DesignQuote).filter(DesignQuote.design_id == design_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    db.delete(quote)
    db.commit()
    return {"message": "Quote deleted"}


def style_header_cell(cell):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_currency_cell(cell, value):
    """Apply currency formatting to a cell."""
    if value is None:
        cell.value = "N/A"
        cell.font = Font(italic=True, color="888888")
    else:
        cell.value = value
        cell.number_format = '"$"#,##0.00'


def _format_currency(value):
    """Format a number as currency string."""
    if value is None:
        return "N/A"
    return f"${value:,.2f}"


def _generate_pdf_export(design: Design, quote: DesignQuote) -> StreamingResponse:
    """Generate PDF export for design with quote."""
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        textColor=colors.HexColor('#1a1a1a')
    )
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#333333')
    )
    normal_style = styles['Normal']

    elements = []

    # Title
    elements.append(Paragraph(f"King Cap - Design #{design.design_number} Quote", title_style))
    elements.append(Spacer(1, 0.2*inch))

    # Design Information Table
    elements.append(Paragraph("Design Information", section_style))
    design_data = [
        ["Design #:", str(design.design_number)],
        ["Design Name:", design.design_name or f"Design #{design.design_number}"],
        ["Brand:", design.brand_name],
        ["Customer:", design.customer_name],
        ["Hat Style:", design.hat_style.replace("-", " ").title()],
        ["Material:", design.material.replace("-", " ").title()],
    ]
    design_table = Table(design_data, colWidths=[1.5*inch, 4*inch])
    design_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(design_table)
    elements.append(Spacer(1, 0.3*inch))

    # Quote Details
    elements.append(Paragraph("Quote Details", section_style))
    quote_data = [
        ["Quote Type:", quote.quote_type.title()],
        ["Quantity:", f"{quote.quantity:,}"],
    ]

    if quote.quote_type == "domestic":
        quote_data.extend([
            ["Style Number:", quote.style_number or "N/A"],
            ["Shipping:", quote.shipping_speed or "Standard"],
        ])
        if quote.include_rope:
            quote_data.append(["Rope:", "Yes"])
    else:
        quote_data.extend([
            ["Hat Type:", quote.hat_type or "N/A"],
            ["Shipping:", quote.shipping_method or "FOB CA"],
        ])

    # Decorations
    if quote.front_decoration:
        quote_data.append(["Front Decoration:", quote.front_decoration])
    if quote.left_decoration:
        quote_data.append(["Left Decoration:", quote.left_decoration])
    if quote.right_decoration:
        quote_data.append(["Right Decoration:", quote.right_decoration])
    if quote.back_decoration:
        quote_data.append(["Back Decoration:", quote.back_decoration])
    if quote.visor_decoration:
        quote_data.append(["Visor Decoration:", quote.visor_decoration])

    quote_table = Table(quote_data, colWidths=[1.5*inch, 4*inch])
    quote_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(quote_table)
    elements.append(Spacer(1, 0.3*inch))

    # Pricing Summary
    elements.append(Paragraph("Pricing Summary", section_style))
    pricing_data = []
    if quote.cached_per_piece:
        pricing_data.append(["Per Piece:", _format_currency(quote.cached_per_piece / 100)])
    if quote.cached_total:
        pricing_data.append(["Total:", _format_currency(quote.cached_total / 100)])

    if pricing_data:
        pricing_table = Table(pricing_data, colWidths=[1.5*inch, 4*inch])
        pricing_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#2563eb')),
        ]))
        elements.append(pricing_table)
    elements.append(Spacer(1, 0.3*inch))

    # Price Breaks Table
    price_breaks = _parse_json_field(quote.cached_price_breaks)
    if price_breaks:
        elements.append(Paragraph("Price Breaks", section_style))
        pb_data = [["Quantity", "Per Piece", "Total"]]
        for pb in price_breaks:
            if pb.get("per_piece_price"):
                total = pb.get("total") or (pb["per_piece_price"] * pb["quantity_break"])
                pb_data.append([
                    f"{pb['quantity_break']:,}+",
                    _format_currency(pb["per_piece_price"]),
                    _format_currency(total)
                ])
            else:
                pb_data.append([f"{pb['quantity_break']:,}+", "Does not meet MOQ", "-"])

        pb_table = Table(pb_data, colWidths=[1.5*inch, 2*inch, 2*inch])
        pb_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(pb_table)

    # Build PDF
    doc.build(elements)
    output.seek(0)

    filename = f"design_{design.design_number}_quote.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{design_id}/quote/export")
async def export_design_with_quote(
    design_id: str,
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    db: Session = Depends(get_db),
    user=Depends(require_auth),
):
    """Export design with quote as Excel or PDF."""
    # Get design and quote
    design = db.query(Design).filter(Design.id == design_id).first()
    if not design:
        raise HTTPException(status_code=404, detail="Design not found")

    quote = db.query(DesignQuote).filter(DesignQuote.design_id == design_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found for this design")

    if format == "pdf":
        return _generate_pdf_export(design, quote)

    # Excel export
    wb = Workbook()
    ws = wb.active
    ws.title = "Design Quote"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"King Cap - Design #{design.design_number} Quote"
    ws["A1"].font = Font(bold=True, size=14)

    # Design info
    row = 3
    design_info = [
        ("Design #:", str(design.design_number)),
        ("Design Name:", design.design_name or f"Design #{design.design_number}"),
        ("Brand:", design.brand_name),
        ("Customer:", design.customer_name),
        ("Hat Style:", design.hat_style.replace("-", " ").title()),
        ("Material:", design.material.replace("-", " ").title()),
    ]
    for label, value in design_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Separator
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Quote Details"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    # Quote info
    quote_info = [
        ("Quote Type:", quote.quote_type.title()),
        ("Quantity:", f"{quote.quantity:,}"),
    ]

    if quote.quote_type == "domestic":
        quote_info.extend([
            ("Style Number:", quote.style_number or "N/A"),
            ("Shipping:", quote.shipping_speed or "Standard"),
        ])
        if quote.include_rope:
            quote_info.append(("Rope:", "Yes"))
    else:
        quote_info.extend([
            ("Hat Type:", quote.hat_type or "N/A"),
            ("Shipping:", quote.shipping_method or "FOB CA"),
        ])

    # Decorations
    if quote.front_decoration:
        quote_info.append(("Front Decoration:", quote.front_decoration))
    if quote.left_decoration:
        quote_info.append(("Left Decoration:", quote.left_decoration))
    if quote.right_decoration:
        quote_info.append(("Right Decoration:", quote.right_decoration))
    if quote.back_decoration:
        quote_info.append(("Back Decoration:", quote.back_decoration))
    if quote.visor_decoration:
        quote_info.append(("Visor Decoration:", quote.visor_decoration))

    for label, value in quote_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Pricing summary
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Pricing Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    if quote.cached_per_piece:
        ws[f"A{row}"] = "Per Piece:"
        ws[f"A{row}"].font = Font(bold=True)
        style_currency_cell(ws[f"B{row}"], quote.cached_per_piece / 100)
        row += 1

    if quote.cached_total:
        ws[f"A{row}"] = "Total:"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        style_currency_cell(ws[f"B{row}"], quote.cached_total / 100)
        ws[f"B{row}"].font = Font(bold=True, size=12)
        row += 1

    # Price breaks table
    price_breaks = _parse_json_field(quote.cached_price_breaks)
    if price_breaks:
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "Price Breaks"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ["Quantity", "Per Piece", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            style_header_cell(cell)
        row += 1

        # Data rows
        for pb in price_breaks:
            ws.cell(row=row, column=1, value=f"{pb['quantity_break']:,}+")
            if pb.get("per_piece_price"):
                style_currency_cell(ws.cell(row=row, column=2), pb["per_piece_price"])
                total = pb.get("total") or (pb["per_piece_price"] * pb["quantity_break"])
                style_currency_cell(ws.cell(row=row, column=3), total)
            else:
                ws.cell(row=row, column=2, value="Does not meet MOQ")
                ws.cell(row=row, column=2).font = Font(italic=True, color="888888")
            row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"design_{design.design_number}_quote.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
